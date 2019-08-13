#!/usr/bin/env python
# coding: utf-8

# In[120]:


import docx2txt
import pycountry
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import time

name = input("Please input file name (i.e. consultant's name)")
text = docx2txt.process("C:\\Users\\Jing Li\\Desktop\\cv2db\\project\\Inputs\\{}.docx".format(name))


# In[45]:


# import os
# os.getcwd()


# In[46]:


# Education
degrees = {"Ph.D.": ["Postdoctoral", "PhD", "Ph.D.", "Doctor of Philosophy", "Doctoral", "Doctor of Education", "Ed.D.",                     "D.D.S.", "DDS", "Doctor of Literature", "Doctor in Public Health",                     "Doctor of Public Health", "Doctor of Sociology"],
           "M.D.": ["M.D.", "Doctor of Medicine", "Doctorate in Medicine"],
           "MBA": ["MBA"],
           "Master": ["Master", "MPH", "M.P.H.", "M.S.", "M.S.N.", "M.A.", "M.P.", "Graduate Studies"],
           "Bachelor": ["Bachelor", "BA", "B.A.","A.B.", "B.S.", "BS", "Undergraduate"]}

def CheckDegree (p_edu):
    p_degree = []
    for i in degrees:
        for title in degrees[i]:
            if title in p_edu:
#             if re.search(title, text, re.IGNORECASE):
                p_degree.append(i)
                break
    return p_degree


# In[47]:


#Language

lang_dic = ['Amharic', 'Arabic', 'Bemba', 'Chinese', 'Creole', 'Danish', 'French', 'German', 'Hindi',            'Kinyarwanda', 'Kirundi', 'Kiswahili', 'Luo', 'Malayalam', 'Nyanja', 'Portuguese', 'Russian',            'Setswana', 'Spanish', 'Swahili', 'Tagalog', 'Tigrigna', 'Tigrinya', 'Yoruba']
lang = []
for i in lang_dic:
    if re.search(i, text, re.IGNORECASE):
        lang.append(i)


# In[48]:


#Countries worked
cnty = []
for country in pycountry.countries:
    if re.search(country.name, text, re.IGNORECASE):
        cnty.append(country.name)


# In[49]:


#Target population
population = {'Early Childhood': ["early child"],
              'Elderly population': ["elder", "senior health"],
              'Orphans': ["Orphan"],
              'Malaria': ["Malaria"],
              'Female population': ["female", "girl", "women", "woman"],
              'HIV/AIDS population': ["hiv", "aids"],
              'Homeless population': ["homeless"],
              'Infants': ["infants", "baby"],
              'LGBTQ': ["lgbt", "transgender"],
              'Low-income population': ["low-income", "poor", "poverty"],
              'Maternal and Children': ["mother", "maternal", "child"],
              'People with disability': ["disable", "disabilit"],
              'Rural Health': ["rural"],
              'Youth': ["youth", "adolescent", "teen"],
              'Tribal Home Visiting': ["tribal"],
             }
pop = []
for k, v in population.items():
    for title in v:
        if re.search(title, text, re.IGNORECASE):
            pop.append(k)
            break


# In[71]:


#Extract Certificate
certificates = {'Advanced Cardiovascular Life Support (ACLS)': ["ACLS"],
                'American Academy of HIV Medicine Specialist (AAHIVS)': ["AAHIVS"],
                'Association of Chartered Certified Accountants (ACCA)': ["ACCA"],
                'Basic Life Support (BLS)': ["BLS"],
                'Cardiopulmonary Resuscitation (CPR)': ["CPR"],
                'Certified Coach, including BCC, PCC, MCC': ["certified coach", "Certified coach", "BCC", "PCC", "MCC"],
                'Certified Grants Management Analyst (CGMS)': ["CGMS"],
                'Certified Professional in Healthcare Quality (CPHQ)': ["CPHQ"],
                'Certified Professional In-Patient Safety (CPPS)': ["CPPS"],
                'Certified Public Accountant (CPA)': ["CPA"],
                'Certified Registered Nurse Practitioner (CRNP)': ["NP", "N.P.", "Nurse Practitioner", "nurse practitioner"],
                'Certified Valuation Analyst (CVA)': ["CVA"],
                'Chartered Global Management Accountant (CGMA)': ["CGMA"],
                'DEA Certificate': ["DEA"],
                'Master Analyst in Financial Forensics (MAFF)': ["MAFF"],
                'Physician Assistant (PA) from NCCPA': ["Physician Assistant", "NCCPA"],
                'Project Management Professional (PMP)': ["PMP"],
                'Registered Nurse (RN)': ["RN", "R.N", "Registered nurse", "Registered Nurse", "registered nurse"],
                'Certified Fraud Examiner (CFE)': ["CFE"],
                'Licensed Physical Therapist': ['PT', 'Physical Therapist']
               }

p_certificate = []
for k, v in certificates.items():
    for i in v:
        if re.search(i, text):
                p_certificate.append(k)
                break


# In[72]:


#Extract Expertises
expertises = {'Coaching & Training': ["coach"],
              'Federal government agency': ["federal"],
              'Meeting Facilitation': ["meeting"],
              'Report Writing': ["report", "proposal", "grant"],
              'Budget and fiscal management': ["budget", "fiscal manage",
                                               "fiscal consul", "fiscal oversight"],
              'Epidemiology': ['Epidemiology'],
              'Global health': ["global health"],
              'Healthcare policy and regulations': ["policy", "regulation"],
              'Infectious Diseases': ["Infectious"],
              'International Relations': ["international relation"],
              'Behavioral Science': ['behavioral'],
              'Mental Health': ['mental health'],
              'Psychiatry': ["Psychiatry"],
              'Substance abuse': ["substance abuse"],
              'Human Trafficking': ["trafficking"],
              'Nutrition/Food security': ["nutrition", "food"],
              'Hepatitis C Virus (HCV)': ["HCV", "Hepatitis C"],
              'Program income reporting': ["program income"],
              'State Compliance Monitoring': ["compliance"],
              'Capacity building': ["capacity building", "capacity develop",
                                    "building their capacity", "build their capacity",
                                   "build capacity", "services capacity", "system capacity",
                                   "capacity expansion", "built capacity"],
              'Clinic Operations': ["clinic"],
              'Financial analysis & management': ["financial", "finance", "accounting"],
              'Fund development & Sustainability': ["fund ", "funds", "funding"],
              'Grants management': ["grants"],
              'Human Capital Systems': ["Human Capital"],
              'Human resources management': ["Human resource"],
              'Humility and Organizational Leaders': ['Humility'],
              'Organizational Development&Transformation': ['Organizational Development',
                                                           'Transformation'],
              'Quality improvement and management': ["quality program", "quality improve",
                                                     "improve quality"],
              'State Procurment System': ["procurement"],
              'Strategic planning': ["Strategic"],
              'Auditing Federally-funded Grants Program': ["Federally-funded"],
              'Federal Acquisition Regulation (FAR)': ["FAR"],
              'Nurse/Administration': ["Nurse"],
              'Data collection, analysis and reporting': ["data collect", "data analy", 'analysis'],
              'Family Planning and Reproductive Health': ["family planning", "FP/RH"],
              'Program evaluation & analysis': ["evaluation", "evalute", "evaluator"],
              'Programmatic Assessment & Management': ["Programmatic"],
              'Time-constrainted Financial Analysis': ["financial"]
             }

e = []
for k, v in expertises.items():
    for i in v:
        if re.search(i, text):
                e.append(k)
                break


# In[73]:


#Extract Federal Project
fp = {'Rural Health (FORHP)': ['FORHP'],
      'Health Centers (BPHC)': ['BPHC'],
      'HIV/AIDS Domestic (HAB Ryan White)': ['Ryan White'],
      'HIV/AIDS International (HAB PEPFAR)': ['PEPFAR'],
      'Maternal and Child Heath (MCHB, ACF)': ['MCHB', "maternal and child"],
      'Maternal, Infant, and Early Childhood Home Visiting Program (MIECHV)': ['MIECHV'],
      'Population Assessment of Tobacco and Health Study (PATH)': ['PATH '],
      'Global Trade Analysis Project (GTAP)': ['GTAP'],
      'The Joint United Nations Programme on HIV/AIDS (UNAIDS)': ['UNAIDS'],
      'Nursing Education Partnership Initiative (NEPI)': ['NEPI'],
      'Head Start': ['Head Start'],
      'Early Head Start': ['Early Head Start', 'EHS'],
      'Community Services Block Grant (CSBG)': ['CSBG'],
      'The HIV/AIDS Bureau (HAB)': ['HAB'],
      'Environmental Protection Agency (EPA)': ['EPA'],
      'General Services Administration (GSA)': ['GSA'],
      'Department of Energy (DOE)': ['DOE'],
      'Government Accountability Office (GAO)': ['GAO'],
      'Statement of Budgetary Resources (SBR)': ['SBR'],
      'Enterprise Resource Programs (ERP)': ['ERP'],
      'Strengthening Accountabilityin the Glocal Economy (SAGE) Fund': ['SAGE'],
      'Head Start Enterprise System': ['Head Start Enterprise System'],
      'Temporary Assistance for Needy Families (TANF)': ['TANF'],
      'Community Health Center (CHC)': ['CHC'],
      'Federal Highway programs': ['Highway pro'],
      'Corporation for National and Community Service (CNCS)': ['CNCS'],
      'Uniform Guidance (2 CFR Part 200)': ['2 CFR Part 200', '2 CFR', ' 2CFR Part 200'],
      'Uniform Guidance (45 CFR Part 75)': ['45CFR75'],
      'Deep Foundations Institute (DFI)': ['DFI'],
      'OMB A-133 Compliance Supplement': ['A-133'],
      'Linking Actions for Unmet Needs in Children’s Health (LAUNCH)': ['LAUNCH'],
      'Healthy Start Initiative': ['Healthy Start'],
      'Home Visiting Collaborative Improvement and Innovation Network (HV CoIIN)': ['CoIIN'],
      'Parents as Teachers (PAT)': ['Parents as Teachers (PAT)'],
      'Race to the Top': ['Race to the Top'],
      'Home Instruction for Parents of Preschool Youngsters (HIPPY)': ['HIPPY'],
      'Healthy Families America (HFA)': ['HFA'],
      'Nurse-Family Partnership (NFP)': ['NFP'],
      'Evidence-based Home Visiting': ['Evidence-based'],
      'Center on the Social and Emotional Foundations for Early Learning (CSEFEL)': ['CSEFEL']
     }

project = []
for k, v in fp.items():
    for i in v:
        if re.search(i, text, re.IGNORECASE):
                project.append(k)
                break


# In[74]:


#Email:
pattern = r"\b[\w\.\d]+\@\w+\.\w+\b"
email = re.search(pattern, text).group(0)


# In[75]:


#Phone
pattern = r"\(?\d{3}\)?.\d{3}[^\d]\d{4}\b"
phone = re.search(pattern, text).group(0)


# In[76]:


#zipcode
pattern = r"\b\d{5}\b"
zipcode = re.search(pattern, text).group(0)


# In[77]:


#state
s = ["Alabama","Alaska","Arizona","Arkansas","California","Colorado",
  "Connecticut","Delaware","Florida","Georgia","Hawaii","Idaho","Illinois",
  "Indiana","Iowa","Kansas","Kentucky","Louisiana","Maine","Maryland",
  "Massachusetts","Michigan","Minnesota","Mississippi","Missouri","Montana",
  "Nebraska","Nevada","New Hampshire","New Jersey","New Mexico","New York",
  "North Carolina","North Dakota","Ohio","Oklahoma","Oregon","Pennsylvania",
  "Rhode Island","South Carolina","South Dakota","Tennessee","Texas","Utah",
  "Vermont","Virginia","Washington","West Virginia","Wisconsin","Wyoming"]

us_state_abbrev = {
    'Alabama': 'AL',
    'Alaska': 'AK',
    'Arizona': 'AZ',
    'Arkansas': 'AR',
    'California': 'CA',
    'Colorado': 'CO',
    'Connecticut': 'CT',
    'Delaware': 'DE',
    'District of Columbia': 'DC',
    'Florida': 'FL',
    'Georgia': 'GA',
    'Hawaii': 'HI',
    'Idaho': 'ID',
    'Illinois': 'IL',
    'Indiana': 'IN',
    'Iowa': 'IA',
    'Kansas': 'KS',
    'Kentucky': 'KY',
    'Louisiana': 'LA',
    'Maine': 'ME',
    'Maryland': 'MD',
    'Massachusetts': 'MA',
    'Michigan': 'MI',
    'Minnesota': 'MN',
    'Mississippi': 'MS',
    'Missouri': 'MO',
    'Montana': 'MT',
    'Nebraska': 'NE',
    'Nevada': 'NV',
    'New Hampshire': 'NH',
    'New Jersey': 'NJ',
    'New Mexico': 'NM',
    'New York': 'NY',
    'North Carolina': 'NC',
    'North Dakota': 'ND',
    'Northern Mariana Islands':'MP',
    'Ohio': 'OH',
    'Oklahoma': 'OK',
    'Oregon': 'OR',
    'Palau': 'PW',
    'Pennsylvania': 'PA',
    'Puerto Rico': 'PR',
    'Rhode Island': 'RI',
    'South Carolina': 'SC',
    'South Dakota': 'SD',
    'Tennessee': 'TN',
    'Texas': 'TX',
    'Utah': 'UT',
    'Vermont': 'VT',
    'Virgin Islands': 'VI',
    'Virginia': 'VA',
    'Washington': 'WA',
    'West Virginia': 'WV',
    'Wisconsin': 'WI',
    'Wyoming': 'WY',
}
pattern =  "|".join(s+list(us_state_abbrev.values()))
state = re.search(pattern, text).group(0)

if state in us_state_abbrev:
    state = us_state_abbrev[state]


# In[111]:


#Country
if state:
    country = "USA"
else:
    country = None


# In[78]:


#city
pattern =  "|".join(s+list(us_state_abbrev.values()))
sta = ", " + re.search(pattern, text).group(0)
pattern = "[\w ]+(?={})".format(sta)
if re.search(pattern, text):
    city = re.search(pattern, text).group(0)
else:
    city = None


# In[79]:


#State Worked
sta_worked = []
for k, v in us_state_abbrev.items():
    if re.search(k, text) or re.search(v, text):
        sta_worked.append(k)


# In[80]:


#Name
f_name = " ".join(name.split(" ")[:-1])
l_name = name.split(" ")[-1]


# In[136]:


#Check if this consultant is in roaster already

wb = load_workbook("C:\\Users\\Jing Li\\Desktop\\cv2db\\project\\Inputs\\master roster.xlsx")
roaster = wb.worksheets[0]
names = []
for row in roaster.iter_rows(min_row=2, min_col = 2, max_col=3):
    names.append(" ".join([cell.value for cell in row][::-1]))
if name in names:
    print("{} has already existed in the roaster!".format(name))
else:
    #Redcord this new consultant
    roaster.append([len(list(roaster.values)) + 1, #c_id
                   l_name, #l_name
                   f_name, #f_name
                   None, #labor_category
                   None, #title
                   None, #category
                   None, #company
                   email, #email
                   phone, #phone_no
                   "Raw Resume", #data_source
                   zipcode, #zipcode
                   None, #address
                   city, #city
                   state, #state
                   country, #country
                   None, #if_interested
                   None, #notes
                   None, #Linkedin
                   "; ".join(project), #Federal Project
                   "; ".join(pop), #Target Population
                   "; ".join(e), #Expertise
                   "; ".join(CheckDegree(text)), #Education
                   "; ".join(p_certificate), #certificate
                   "; ".join(lang), #Language
                   "; ".join(cnty), #Country Worked
                   "; ".join(sta_worked), #State Worked
                   None, #Site Visited
                   ])
    #Adjust Cell style
    roaster['A{}'.format(len(list(roaster.values)) )].font = Font(bold=True)
    roaster['A{}'.format(len(list(roaster.values)) )].alignment = Alignment(horizontal='center')
    #Save
    wb.save("C:\\Users\\Jing Li\\Desktop\\cv2db\\project\\Inputs\\master roster.xlsx")
    print("{} has been successfully recorded in the roaster!".format(name))

