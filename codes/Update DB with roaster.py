import docx2txt
import pycountry
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import time

# Check difference
wb = load_workbook("C:\\Users\\Jing Li\\Desktop\\cv2db\\project\\Inputs\\master roster.xlsx")
roaster = wb.worksheets[0]
r_ids = []
for row in roaster.iter_rows(min_row=2, min_col = 1, max_col=1):
    if row[0].value:
        r_ids.append(row[0].value)

wb_db = load_workbook("C:\\Users\\Jing Li\\Desktop\\cv2db\\project\\Inputs\\Consultants Database.xlsx")
ws_consultant = wb_db["Consultants"]
db_ids = []
for row in ws_consultant.iter_rows(min_row=2, min_col = 1, max_col=1):
    db_ids.append(row[0].value)

if set(r_ids) == set(db_ids):
    id_update = input("All consultants in roaster have been in database. Please input the id of whom you want to update.")
    diff = [int(id_update)]
else:
    diff = list(set(r_ids) - set(db_ids))
    print("consultant(s) with id{} not in database".format(diff))
    
print("Updating ...")    
# Time Delay for 1 second
time.sleep(1)

#Functions

# Education
degrees = {"Ph.D.": ["Postdoctoral", "PhD", "Ph.D.", "Doctor of Philosophy", "Doctoral", "Doctor of Education", "Ed.D.",                     "D.D.S.", "DDS", "Doctor of Literature", "Doctor in Public Health",                     "Doctor of Public Health", "Doctor of Sociology"],
           "M.D.": ["M.D.", "Doctor of Medicine", "Doctorate in Medicine"],
           "MBA": ["MBA"],
           "Master": ["Master", "MPH", "M.P.H.", "M.S.", "M.S.N.", "M.A.", "M.P.", "Graduate Studies"],
           "Bachelor": ["Bachelor", "BA", "B.A.","A.B.", "B.S.", "BS", "Undergraduate"]}

def check_deg (p_edu):
    for i in degrees:
        for title in degrees[i]:
            if title in p_edu:
                return i
    return None
    
def check_yr(edu):
    if re.search(r'\b\d{4}\b', edu):
        return re.search(r'\b\d{4}\b', edu).group(0)
    return None

#Load worksheet
ws_fp = wb_db["Federal Project"]
ws_tp = wb_db["Target Population"]
ws_e = wb_db["Expertise"]
ws_edu = wb_db["Education"]
ws_c = wb_db["Certificate"]
ws_cul = wb_db["Cultural competency"]

# Read Data
for i in diff: # i is the id of consultant to be updated
    if i >= 20 and i <= len(list(roaster.values)):
        row = i
    elif i <= 18 and i >= 1:
        row = i + 1
    else:
        print("id input is invalid. Please check roaster and your input")


    #Update    
    for row in roaster.iter_rows(min_row=row, max_row=row):
       
        #basic info
        ids = [i[0].value for i in ws_consultant.iter_rows(max_col = 1)]
        if i in ids:
            row_no= ids.index(i) + 1
            ws_consultant.delete_rows(row_no)
            ws_consultant.insert_rows(row_no)
            for j in range(1, 19):
                ws_consultant.cell(row=row_no, column=j, value=row[j-1].value)
        else:
            ws_consultant.append([r.value for r in row[:18]])
        
        #federal project        
        ids = [k[0].value for k in ws_fp.iter_rows(max_col = 1)]
        i_ind = [index for index, x in enumerate(ids) if x == i]
        for I_ind in i_ind:
            ws_fp.delete_rows(i_ind[0] + 1)
        if row[18].value:
            for fp in row[18].value.split("; "):
                ws_fp.append([i, fp])
            
        #target population
        ids = [k[0].value for k in ws_tp.iter_rows(max_col = 1)]
        i_ind = [index for index, x in enumerate(ids) if x == i]
        for I_ind in i_ind:
            ws_tp.delete_rows(i_ind[0]+1)
        if row[19].value:
            for tp in row[19].value.split("; "):
                ws_tp.append([i, tp])
            
        #Expertise
        ids = [k[0].value for k in ws_e.iter_rows(max_col = 1)]
        i_ind = [index for index, x in enumerate(ids) if x == i]
        for I_ind in i_ind:
            ws_e.delete_rows(i_ind[0]+1)
        if row[20].value:
            for e in row[20].value.split("; "):
                ws_e.append([i, e])
            
        #Education
        ids = [k[0].value for k in ws_edu.iter_rows(max_col = 1)]
        i_ind = [index for index, x in enumerate(ids) if x == i]
        for I_ind in i_ind:
            ws_edu.delete_rows(i_ind[0]+1)
        if row[21].value:
            for edu in row[21].value.split("; "):
                deg = check_deg(edu)
                yr = check_yr(edu)
                ws_edu.append([i, deg, edu, yr])
            
        #Certificate
        ids = [k[0].value for k in ws_c.iter_rows(max_col = 1)]
        i_ind = [index for index, x in enumerate(ids) if x == i]
        for I_ind in i_ind:
            ws_c.delete_rows(i_ind[0]+1)
        if row[22].value:
            for c in row[22].value.split("; "):
                ws_c.append([i, c])     
            
        #Cultural competency
        #lANGUAGE
        ids = [k[0].value for k in ws_cul.iter_rows(max_col = 1)]
        i_ind = [index for index, x in enumerate(ids) if x == i]
        for I_ind in i_ind:
            ws_cul.delete_rows(i_ind[0]+1)
        if row[23].value:
            for cul in row[23].value.split("; "):
                ws_cul.append([i, "Language", cul])
            #Adjust Cell style
                ws_cul['A{}'.format(len(list(ws_cul.values)) )].font = Font(bold=True)
                ws_cul['A{}'.format(len(list(ws_cul.values)) )].alignment = Alignment(horizontal='center')
                
        #COUNTRY WORKED
        if row[24].value:
            for cul in row[24].value.split("; "):
                ws_cul.append([i, "Country", cul])
            #Adjust Cell style
                ws_cul['A{}'.format(len(list(ws_cul.values)) )].font = Font(bold=True)
                ws_cul['A{}'.format(len(list(ws_cul.values)) )].alignment = Alignment(horizontal='center')
                
        #STATE WORKED
        if row[25].value:
            for cul in row[25].value.split("; "):
                ws_cul.append([i, "State", cul])
            #Adjust Cell style
                ws_cul['A{}'.format(len(list(ws_cul.values)) )].font = Font(bold=True)
                ws_cul['A{}'.format(len(list(ws_cul.values)) )].alignment = Alignment(horizontal='center')
                
        #SITE VISITED
        if row[26].value:
            for cul in row[26].value.split("; "):
                ws_cul.append([i, "Site Visit", cul])
            #Adjust Cell style
                ws_cul['A{}'.format(len(list(ws_cul.values)) )].font = Font(bold=True)
                ws_cul['A{}'.format(len(list(ws_cul.values)) )].alignment = Alignment(horizontal='center')    

#Save
wb_db.save("C:\\Users\\Jing Li\\Desktop\\cv2db\\project\\Inputs\\Consultants Database.xlsx")
print("Consultant(s) with id{} updated finished :D".format(diff))